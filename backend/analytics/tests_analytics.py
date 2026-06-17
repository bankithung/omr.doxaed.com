"""
TDD tests for Phase 5, Task 1: test-level analytics endpoint (shuffle-correct).

Tests cover:
1. QuestionResponse.question is populated by the pipeline grading tweak.
2. test_summary() computations: average, median, distribution, toppers.
3. A question everyone got wrong tops hardest_questions.
4. option_distribution maps printed→original labels correctly under an option shuffle.
5. GET /api/v1/analytics/test/{test_id}/ endpoint (auth, 200, shape, cross-tenant 404).
"""

from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APIClient

from assessments.models import ClassGroup, MarkingScheme, Option, Question, Test
from omr.models import OmrSheet
from results.models import QuestionResponse, StudentResult
from rosters.models import Roster, Student

User = get_user_model()


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _make_user(email="teacher@test.com", password="pass1234"):
    return User.objects.create_user(email=email, password=password)


def _make_test(user, title="Analytics Test"):
    cg = ClassGroup.objects.create(user=user, created_by=user, name="CG-Analytics")
    t = Test.objects.create(
        user=user, created_by=user, class_group=cg,
        title=title, subject="Math",
    )
    MarkingScheme.objects.create(test=t, marks_per_correct=1)
    return t


def _make_questions(test, n=3):
    """Create n questions each with 4 options (A correct)."""
    questions = []
    for i in range(n):
        q = Question.objects.create(
            test=test,
            order_index=i,
            text=f"Question {i+1}",
            topic=f"Topic {(i % 2) + 1}",
        )
        for label in ["A", "B", "C", "D"]:
            Option.objects.create(
                question=q, label=label, text=label, is_correct=(label == "A")
            )
        questions.append(q)
    return questions


def _make_student(user, roll="001", name="Student One"):
    roster, _ = Roster.objects.get_or_create(user=user, created_by=user, name="Roster-A")
    return Student.objects.create(roster=roster, roll_number=roll, full_name=name)


def _make_omr_sheet(test, questions, seed=1):
    """Build an OmrSheet with identity shuffle (question_order and option_order)."""
    q_ids = [q.id for q in questions]
    option_order = {str(q.id): ["A", "B", "C", "D"] for q in questions}
    # answer_key: printed position → [correct printed labels] (same as original, no shuffle)
    answer_key = {str(i): ["A"] for i in range(len(questions))}
    return OmrSheet.objects.create(
        test=test,
        sheet_code=f"SHEET-{seed:04d}",
        human_readable_code=f"S{seed:04d}",
        question_order=q_ids,
        option_order=option_order,
        answer_key=answer_key,
        page_count=1,
        shuffle_version=seed,
    )


def _make_omr_sheet_shuffled(test, questions, seed=99):
    """
    Build an OmrSheet with a KNOWN option shuffle on question[0]:
        question[0] original labels: A B C D
        shuffled (printed) order:    B A D C
    So printed label 'A' (printed pos 1) corresponds to original 'B'.
    And printed label 'B' (printed pos 0) corresponds to original 'A' (correct).
    """
    q_ids = [q.id for q in questions]
    q0_id = str(questions[0].id)
    # question[0]: printed order is [B, A, D, C]
    # question[1], question[2]: identity
    option_order = {str(q.id): ["A", "B", "C", "D"] for q in questions}
    option_order[q0_id] = ["B", "A", "D", "C"]
    # answer key: pos 0 → correct printed label is 'B' (maps to original 'A')
    # for positions 1 and 2: 'A' is still correct (identity shuffle)
    answer_key = {"0": ["B"], "1": ["A"], "2": ["A"]}
    return OmrSheet.objects.create(
        test=test,
        sheet_code=f"SHEET-SHF-{seed:04d}",
        human_readable_code=f"SS{seed:04d}",
        question_order=q_ids,
        option_order=option_order,
        answer_key=answer_key,
        page_count=1,
        shuffle_version=seed,
    )


def _make_result(test, student, omr_sheet, score, max_score=3,
                 correct=0, wrong=0, blank=0, needs_review=False):
    return StudentResult.objects.create(
        test=test,
        student=student,
        omr_sheet=omr_sheet,
        score=Decimal(str(score)),
        max_score=Decimal(str(max_score)),
        correct_count=correct,
        wrong_count=wrong,
        blank_count=blank,
        needs_review=needs_review,
    )


def _make_qr(result, question, q_pos, marked, is_correct, flagged=False):
    return QuestionResponse.objects.create(
        student_result=result,
        question=question,
        q_pos=q_pos,
        marked_options=marked,
        is_correct=is_correct,
        flagged=flagged,
    )


# ===========================================================================
# 1. Pipeline grading tweak: QuestionResponse.question FK is populated
# ===========================================================================

class PipelineQuestionFKTest(TestCase):
    """
    Test that _maybe_grade sets QuestionResponse.question from
    omr_sheet.question_order[q_pos].

    We build a minimal DB fixture and call _maybe_grade via the pipeline
    helper simulate_correct_marks + grade_sheet, then directly call
    _maybe_grade to avoid the full image pipeline.
    """

    def _build_fixtures(self):
        user = _make_user("pipeline@test.com")
        test = _make_test(user)
        questions = _make_questions(test, n=3)
        student = _make_student(user, "P001", "Pipeline Student")
        sheet = _make_omr_sheet(test, questions, seed=500)
        sheet.student = student
        sheet.save()
        return user, test, questions, student, sheet

    def test_question_fk_set_on_grading(self):
        """
        After _maybe_grade runs, each QuestionResponse.question_id must
        equal the underlying Question id from omr_sheet.question_order[q_pos].
        """
        from omr.models import ScanBatch, ScanJob
        from omr.scan.pipeline import _maybe_grade

        user, test, questions, student, sheet = self._build_fixtures()

        # Create a ScanBatch + ScanJob with pre-filled reads (all correct)
        batch = ScanBatch.objects.create(test=test, created_by=user, total=1)
        job = ScanJob.objects.create(
            batch=batch,
            omr_sheet=sheet,
            page_no=1,
            status=ScanJob.STATUS_DONE,
            reads={
                "answers": {
                    "0": {"marked": ["A"], "flag": None},
                    "1": {"marked": ["A"], "flag": None},
                    "2": {"marked": ["A"], "flag": None},
                },
                "roll": "P001",
                "flags": [],
            },
        )

        _maybe_grade(sheet)

        result = StudentResult.objects.get(omr_sheet=sheet)
        responses = list(result.responses.order_by("q_pos"))
        self.assertEqual(len(responses), 3)

        for resp in responses:
            q_pos = resp.q_pos
            expected_question_id = sheet.question_order[q_pos]
            self.assertIsNotNone(
                resp.question_id,
                f"QuestionResponse at q_pos={q_pos} should have question set",
            )
            self.assertEqual(
                resp.question_id,
                expected_question_id,
                f"q_pos={q_pos}: expected question_id={expected_question_id}, "
                f"got {resp.question_id}",
            )


# ===========================================================================
# 2. test_summary() pure-function tests
# ===========================================================================

class TestSummaryServiceTest(TestCase):
    """Tests for analytics.services.test_summary(test)."""

    def setUp(self):
        self.user = _make_user("summary@test.com")
        self.test = _make_test(self.user)
        self.questions = _make_questions(self.test, n=3)

    def _build_scenario(self):
        """
        Create 4 students with scores 3, 2, 1, 0 out of 3:
          - student A: score=3, all correct
          - student B: score=2, q0 wrong, q1 correct, q2 correct
          - student C: score=1, q0 wrong, q1 wrong, q2 correct
          - student D: score=0, all wrong
        """
        students = []
        sheets = []
        results = []
        for i, (score, name, roll) in enumerate([
            (3, "Alice", "001"),
            (2, "Bob", "002"),
            (1, "Carol", "003"),
            (0, "Dan", "004"),
        ]):
            s = _make_student(self.user, roll, name)
            sheet = _make_omr_sheet(self.test, self.questions, seed=i + 10)
            sheet.student = s
            sheet.save()
            correct = score
            wrong = 3 - score
            r = _make_result(
                self.test, s, sheet, score, max_score=3,
                correct=correct, wrong=wrong, blank=0,
            )
            # QuestionResponse rows — all correct up to `score` questions
            for q_idx, q in enumerate(self.questions):
                is_correct = (q_idx < score)
                marked = ["A"] if is_correct else ["B"]
                _make_qr(r, q, q_idx, marked, is_correct)
            students.append(s)
            sheets.append(sheet)
            results.append(r)
        return students, sheets, results

    def test_n_students_and_graded(self):
        from analytics.services import test_summary
        self._build_scenario()
        summary = test_summary(self.test)
        self.assertEqual(summary["n_students"], 4)
        self.assertEqual(summary["graded"], 4)

    def test_needs_review_count(self):
        from analytics.services import test_summary
        students, sheets, results = self._build_scenario()
        # Mark first result as needs_review
        results[0].needs_review = True
        results[0].save()
        summary = test_summary(self.test)
        self.assertEqual(summary["needs_review_count"], 1)

    def test_average(self):
        from analytics.services import test_summary
        self._build_scenario()
        # scores: 3, 2, 1, 0 → average = 6/4 = 1.5
        summary = test_summary(self.test)
        self.assertAlmostEqual(float(summary["average"]), 1.5, places=4)

    def test_median(self):
        from analytics.services import test_summary
        self._build_scenario()
        # sorted scores: 0, 1, 2, 3 → median = (1+2)/2 = 1.5
        summary = test_summary(self.test)
        self.assertAlmostEqual(float(summary["median"]), 1.5, places=4)

    def test_max_and_min(self):
        from analytics.services import test_summary
        self._build_scenario()
        summary = test_summary(self.test)
        self.assertEqual(float(summary["max"]), 3.0)
        self.assertEqual(float(summary["min"]), 0.0)

    def test_max_score(self):
        from analytics.services import test_summary
        self._build_scenario()
        summary = test_summary(self.test)
        self.assertEqual(float(summary["max_score"]), 3.0)

    def test_distribution_bucket_counts(self):
        """
        Scores: 3, 2, 1, 0 with max=3.
        Percentages: 100%, 66.7%, 33.3%, 0%
        Expected buckets:
          0-20%:  1 (score=0 → 0%)
          21-40%: 1 (score=1 → 33.3%)
          41-60%: 0
          61-80%: 1 (score=2 → 66.7%)
          81-100%: 1 (score=3 → 100%)
        """
        from analytics.services import test_summary
        self._build_scenario()
        summary = test_summary(self.test)
        dist = {b["bucket"]: b["count"] for b in summary["distribution"]}
        self.assertEqual(dist["0-20%"], 1)
        self.assertEqual(dist["21-40%"], 1)
        self.assertEqual(dist["41-60%"], 0)
        self.assertEqual(dist["61-80%"], 1)
        self.assertEqual(dist["81-100%"], 1)

    def test_toppers_top5_by_score(self):
        """toppers should be sorted descending by score, max 5 entries."""
        from analytics.services import test_summary
        self._build_scenario()
        summary = test_summary(self.test)
        toppers = summary["toppers"]
        self.assertLessEqual(len(toppers), 5)
        # First topper should have score=3
        self.assertEqual(float(toppers[0]["score"]), 3.0)
        # Ensure descending order
        scores = [float(t["score"]) for t in toppers]
        self.assertEqual(scores, sorted(scores, reverse=True))

    def test_toppers_include_roll_and_name(self):
        """Each topper entry must include student roll and name."""
        from analytics.services import test_summary
        self._build_scenario()
        summary = test_summary(self.test)
        for t in summary["toppers"]:
            self.assertIn("student", t)
            self.assertIn("roll", t["student"])
            self.assertIn("name", t["student"])

    def test_hardest_questions_most_missed_is_first(self):
        """
        The question everyone got wrong (Dan got all wrong, Carol got 2 wrong,
        Bob got 1 wrong) is question[0].
        wrong_rate for q[0]: wrong=3 (Alice+Carol+Dan); n=4 → 3/4=0.75
        wrong_rate for q[1]: wrong=2 (Carol+Dan);       n=4 → 2/4=0.50
        wrong_rate for q[2]: wrong=1 (Dan);              n=4 → 1/4=0.25
        """
        from analytics.services import test_summary
        self._build_scenario()
        summary = test_summary(self.test)
        hq = summary["hardest_questions"]
        self.assertGreater(len(hq), 0)
        # Top entry should be the most-missed question
        self.assertAlmostEqual(float(hq[0]["wrong_rate"]), 0.75, places=4)
        # Ensure descending order
        rates = [float(h["wrong_rate"]) for h in hq]
        self.assertEqual(rates, sorted(rates, reverse=True))

    def test_hardest_questions_include_text_and_order_index(self):
        from analytics.services import test_summary
        self._build_scenario()
        summary = test_summary(self.test)
        for h in summary["hardest_questions"]:
            self.assertIn("question_id", h)
            self.assertIn("text", h)
            self.assertIn("order_index", h)
            self.assertIn("wrong_rate", h)
            self.assertIn("n", h)

    def test_empty_test_no_results(self):
        """test_summary on a test with no StudentResults returns sane defaults."""
        from analytics.services import test_summary
        summary = test_summary(self.test)
        self.assertEqual(summary["n_students"], 0)
        self.assertEqual(summary["graded"], 0)
        self.assertIsNone(summary["average"])
        self.assertIsNone(summary["median"])
        self.assertIsNone(summary["max"])
        self.assertIsNone(summary["min"])
        self.assertEqual(summary["toppers"], [])
        self.assertEqual(summary["hardest_questions"], [])

    def test_distribution_has_five_buckets(self):
        from analytics.services import test_summary
        self._build_scenario()
        summary = test_summary(self.test)
        self.assertEqual(len(summary["distribution"]), 5)
        bucket_names = [b["bucket"] for b in summary["distribution"]]
        self.assertEqual(
            bucket_names,
            ["0-20%", "21-40%", "41-60%", "61-80%", "81-100%"],
        )


# ===========================================================================
# 3. option_distribution: printed→original mapping under option shuffle
# ===========================================================================

class OptionDistributionTest(TestCase):
    """
    Verifies that option_distribution correctly maps printed marked labels
    to original labels via option_order, even when options are shuffled.

    Sheet layout for student 1 (SHUFFLED):
        question[0]: option_order = ["B", "A", "D", "C"]
            → printed 'A' == original 'B', printed 'B' == original 'A' (correct)
        question[1], [2]: identity (no shuffle)

    Student marks printed 'A' on question[0] — which maps to original 'B' (wrong).
    option_distribution['A'] for q[0] should be 0 (no one marked original 'A'),
    option_distribution['B'] for q[0] should be 1 (one person's 'printed A' = original 'B').
    """

    def setUp(self):
        self.user = _make_user("optdist@test.com")
        self.test = _make_test(self.user)
        self.questions = _make_questions(self.test, n=3)

    def _build_shuffled_scenario(self):
        student = _make_student(self.user, "S01", "Shuffled Student")
        sheet = _make_omr_sheet_shuffled(self.test, self.questions, seed=200)
        sheet.student = student
        sheet.save()
        result = _make_result(
            self.test, student, sheet,
            score=0, max_score=3, correct=0, wrong=3,
        )
        # Student marks PRINTED 'A' on q[0] (which is original 'B' — wrong)
        # Student marks PRINTED 'A' on q[1] (which is original 'A' — correct, identity)
        # Student marks PRINTED 'A' on q[2] (which is original 'A' — correct, identity)
        _make_qr(result, self.questions[0], 0, ["A"], is_correct=False)  # printed A → original B
        _make_qr(result, self.questions[1], 1, ["A"], is_correct=True)   # identity
        _make_qr(result, self.questions[2], 2, ["A"], is_correct=True)   # identity
        return student, sheet, result

    def test_option_distribution_maps_printed_to_original(self):
        from analytics.services import test_summary
        self._build_shuffled_scenario()
        summary = test_summary(self.test)
        opt_dist = summary["option_distribution"]

        # Find entry for question[0]
        q0_id = self.questions[0].id
        q0_entry = next((e for e in opt_dist if e["question_id"] == q0_id), None)
        self.assertIsNotNone(q0_entry, "option_distribution should have entry for question[0]")

        counts_by_label = {o["label"]: o["count"] for o in q0_entry["options"]}
        # Printed 'A' on q[0] in shuffled sheet → original 'B'
        self.assertEqual(
            counts_by_label.get("B", 0), 1,
            f"Original 'B' should have count=1 (one student's printed 'A' maps to it). "
            f"counts: {counts_by_label}",
        )
        # Original 'A' (correct) should have count=0 (nobody marked printed 'B')
        self.assertEqual(
            counts_by_label.get("A", 0), 0,
            f"Original 'A' (correct) should have count=0. counts: {counts_by_label}",
        )

    def test_option_distribution_includes_correct_labels(self):
        """Each option_distribution entry must have the correct original labels."""
        from analytics.services import test_summary
        self._build_shuffled_scenario()
        summary = test_summary(self.test)
        opt_dist = summary["option_distribution"]

        q0_id = self.questions[0].id
        q0_entry = next(e for e in opt_dist if e["question_id"] == q0_id)
        # Correct original label for q[0] is 'A'
        self.assertIn("A", q0_entry["correct"], f"correct labels should include 'A', got {q0_entry['correct']}")

    def test_option_distribution_identity_shuffle(self):
        """With identity shuffle, printed 'A' correctly counts as original 'A'."""
        from analytics.services import test_summary
        student = _make_student(self.user, "I01", "Identity Student")
        sheet = _make_omr_sheet(self.test, self.questions, seed=300)
        sheet.student = student
        sheet.save()
        result = _make_result(self.test, student, sheet, score=3, max_score=3, correct=3)
        for q_idx, q in enumerate(self.questions):
            _make_qr(result, q, q_idx, ["A"], is_correct=True)

        summary = test_summary(self.test)
        opt_dist = summary["option_distribution"]
        for q in self.questions:
            entry = next(e for e in opt_dist if e["question_id"] == q.id)
            counts = {o["label"]: o["count"] for o in entry["options"]}
            self.assertEqual(counts.get("A", 0), 1, f"q={q.id}: original 'A' count should be 1")


# ===========================================================================
# 4. API endpoint tests
# ===========================================================================

class AnalyticsEndpointTest(TestCase):
    """GET /api/v1/analytics/test/{test_id}/ endpoint tests."""

    def setUp(self):
        self.client = APIClient()
        self.user = _make_user("api@test.com")
        self.test = _make_test(self.user)
        self.questions = _make_questions(self.test, n=3)

    def _auth(self, user=None):
        if user is None:
            user = self.user
        self.client.force_authenticate(user=user)

    def _url(self, test_id=None):
        if test_id is None:
            test_id = self.test.id
        return f"/api/v1/analytics/test/{test_id}/"

    def test_unauthenticated_returns_401(self):
        response = self.client.get(self._url())
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_cross_tenant_returns_404(self):
        """A test belonging to another user must return 404."""
        other_user = _make_user("other@test.com")
        other_test = _make_test(other_user)
        self._auth()
        response = self.client.get(self._url(other_test.id))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_nonexistent_test_returns_404(self):
        self._auth()
        response = self.client.get(self._url(99999))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_authenticated_owner_returns_200(self):
        self._auth()
        response = self.client.get(self._url())
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_response_has_required_top_level_keys(self):
        self._auth()
        response = self.client.get(self._url())
        data = response.json()
        for key in ("test", "n_students", "graded", "needs_review_count",
                    "average", "median", "max", "min", "max_score",
                    "distribution", "toppers", "hardest_questions",
                    "option_distribution"):
            self.assertIn(key, data, f"Top-level key {key!r} missing from response")

    def test_response_test_block(self):
        """The 'test' block should contain id, title, subject, attempt_number."""
        self._auth()
        response = self.client.get(self._url())
        data = response.json()
        t_block = data["test"]
        for key in ("id", "title", "subject", "attempt_number"):
            self.assertIn(key, t_block, f"test block missing key {key!r}")
        self.assertEqual(t_block["id"], self.test.id)
        self.assertEqual(t_block["title"], self.test.title)

    def test_response_with_results(self):
        """End-to-end: populated test returns correct average."""
        # Create two students with scores 3 and 1
        for i, (score, roll, name) in enumerate([(3, "T01", "Top"), (1, "T02", "Low")]):
            s = _make_student(self.user, roll, name)
            sheet = _make_omr_sheet(self.test, self.questions, seed=i + 100)
            sheet.student = s
            sheet.save()
            r = _make_result(self.test, s, sheet, score, max_score=3)
            for q_idx, q in enumerate(self.questions):
                is_correct = q_idx < score
                _make_qr(r, q, q_idx, ["A"] if is_correct else ["B"], is_correct)

        self._auth()
        response = self.client.get(self._url())
        data = response.json()
        self.assertEqual(data["n_students"], 2)
        # average = (3+1)/2 = 2.0
        self.assertAlmostEqual(float(data["average"]), 2.0, places=2)
        self.assertEqual(len(data["distribution"]), 5)
        self.assertLessEqual(len(data["toppers"]), 5)

    def test_distribution_is_list_of_dicts_with_bucket_count(self):
        self._auth()
        response = self.client.get(self._url())
        data = response.json()
        for item in data["distribution"]:
            self.assertIn("bucket", item)
            self.assertIn("count", item)

    def test_toppers_max_5(self):
        """Even with 10 students, toppers list is capped at 5."""
        for i in range(10):
            s = _make_student(self.user, f"{i+10:03d}", f"S{i}")
            sheet = _make_omr_sheet(self.test, self.questions, seed=i + 200)
            sheet.student = s
            sheet.save()
            r = _make_result(self.test, s, sheet, i % 4, max_score=3)
            for q_idx, q in enumerate(self.questions):
                _make_qr(r, q, q_idx, ["A"], is_correct=True)

        self._auth()
        response = self.client.get(self._url())
        data = response.json()
        self.assertLessEqual(len(data["toppers"]), 5)


# ===========================================================================
# Task 2 — student_detail() service tests
# ===========================================================================

class StudentDetailServiceTest(TestCase):
    """Tests for analytics.services.student_detail(student_result)."""

    def setUp(self):
        self.user = _make_user("detail@test.com")
        self.test = _make_test(self.user)
        # Three questions: q0 topic=TopicA, q1 topic=TopicA, q2 topic=TopicB
        self.q0 = Question.objects.create(
            test=self.test, order_index=0, text="Q0", topic="TopicA"
        )
        self.q1 = Question.objects.create(
            test=self.test, order_index=1, text="Q1", topic="TopicA"
        )
        self.q2 = Question.objects.create(
            test=self.test, order_index=2, text="Q2", topic="TopicB"
        )
        for q in [self.q0, self.q1, self.q2]:
            for label in ["A", "B", "C", "D"]:
                Option.objects.create(question=q, label=label, text=label, is_correct=(label == "A"))

    def _build_result(self, score=2, max_score=3):
        """
        Student gets q0 correct (A), q1 wrong (B), q2 blank ([]).
        score=2 because: +1 correct, no negative marking.
        Actually for test: q0 correct, q1 wrong, q2 blank → score depends on scheme.
        We set score/correct/wrong/blank directly.
        """
        student = _make_student(self.user, "D001", "Detail Student")
        sheet = _make_omr_sheet(self.test, [self.q0, self.q1, self.q2], seed=700)
        sheet.student = student
        sheet.save()
        result = _make_result(
            self.test, student, sheet,
            score=score, max_score=max_score,
            correct=1, wrong=1, blank=1,
        )
        # q0: correct
        _make_qr(result, self.q0, 0, ["A"], is_correct=True, flagged=False)
        # q1: wrong
        _make_qr(result, self.q1, 1, ["B"], is_correct=False, flagged=True)
        # q2: blank
        _make_qr(result, self.q2, 2, [], is_correct=False, flagged=False)
        return student, sheet, result

    def test_basic_fields(self):
        from analytics.services import student_detail
        _, _, result = self._build_result(score=1, max_score=3)
        detail = student_detail(result)
        self.assertAlmostEqual(detail["score"], 1.0, places=4)
        self.assertAlmostEqual(detail["max_score"], 3.0, places=4)
        self.assertAlmostEqual(detail["pct"], 1.0 / 3.0 * 100.0, places=2)
        self.assertEqual(detail["correct"], 1)
        self.assertEqual(detail["wrong"], 1)
        self.assertEqual(detail["blank"], 1)

    def test_pct_guard_zero_max_score(self):
        from analytics.services import student_detail
        student = _make_student(self.user, "D002", "Zero Max")
        sheet = _make_omr_sheet(self.test, [self.q0, self.q1, self.q2], seed=701)
        sheet.student = student
        sheet.save()
        result = _make_result(self.test, student, sheet, score=0, max_score=0)
        detail = student_detail(result)
        self.assertEqual(detail["pct"], 0.0)

    def test_per_question_count_and_fields(self):
        from analytics.services import student_detail
        _, _, result = self._build_result(score=1, max_score=3)
        detail = student_detail(result)
        pq = detail["per_question"]
        self.assertEqual(len(pq), 3)
        # All required keys
        for entry in pq:
            self.assertIn("order_index", entry)
            self.assertIn("text", entry)
            self.assertIn("is_correct", entry)
            self.assertIn("flagged", entry)
            self.assertIn("marked_options", entry)

    def test_per_question_ordered_by_order_index(self):
        from analytics.services import student_detail
        _, _, result = self._build_result(score=1, max_score=3)
        detail = student_detail(result)
        indices = [e["order_index"] for e in detail["per_question"]]
        self.assertEqual(indices, sorted(indices))

    def test_per_question_correct_and_wrong(self):
        from analytics.services import student_detail
        _, _, result = self._build_result(score=1, max_score=3)
        detail = student_detail(result)
        pq = detail["per_question"]
        # q0 correct, q1 wrong/flagged, q2 blank
        q0_entry = next(e for e in pq if e["order_index"] == 0)
        q1_entry = next(e for e in pq if e["order_index"] == 1)
        q2_entry = next(e for e in pq if e["order_index"] == 2)
        self.assertTrue(q0_entry["is_correct"])
        self.assertFalse(q1_entry["is_correct"])
        self.assertTrue(q1_entry["flagged"])
        self.assertEqual(q0_entry["marked_options"], ["A"])
        self.assertEqual(q1_entry["marked_options"], ["B"])
        self.assertEqual(q2_entry["marked_options"], [])

    def test_topic_accuracy_two_topics(self):
        """
        TopicA has q0 (correct) + q1 (wrong) → 1/2 = 0.5
        TopicB has q2 (blank, not correct) → 0/1 = 0.0
        """
        from analytics.services import student_detail
        _, _, result = self._build_result(score=1, max_score=3)
        detail = student_detail(result)
        ta = detail["topic_accuracy"]
        self.assertIn("TopicA", ta)
        self.assertIn("TopicB", ta)
        self.assertAlmostEqual(ta["TopicA"]["accuracy"], 0.5, places=4)
        self.assertAlmostEqual(ta["TopicB"]["accuracy"], 0.0, places=4)
        self.assertEqual(ta["TopicA"]["correct"], 1)
        self.assertEqual(ta["TopicA"]["total"], 2)
        self.assertEqual(ta["TopicB"]["correct"], 0)
        self.assertEqual(ta["TopicB"]["total"], 1)

    def test_topic_accuracy_untagged_bucket(self):
        """Questions with blank topic fall into 'Untagged'."""
        from analytics.services import student_detail
        # Create a question with no topic
        q_notag = Question.objects.create(
            test=self.test, order_index=3, text="NoTag", topic=""
        )
        Option.objects.create(question=q_notag, label="A", text="A", is_correct=True)

        student = _make_student(self.user, "D003", "Untagged Student")
        sheet = _make_omr_sheet(self.test, [self.q0, self.q1, self.q2, q_notag], seed=702)
        sheet.student = student
        sheet.save()
        result = _make_result(self.test, student, sheet, score=1, max_score=4, correct=1, wrong=0, blank=3)
        _make_qr(result, self.q0, 0, ["A"], is_correct=True)
        _make_qr(result, self.q1, 1, [], is_correct=False)
        _make_qr(result, self.q2, 2, [], is_correct=False)
        _make_qr(result, q_notag, 3, ["A"], is_correct=True)

        detail = student_detail(result)
        ta = detail["topic_accuracy"]
        self.assertIn("Untagged", ta, f"Expected 'Untagged' key, got: {list(ta.keys())}")
        self.assertAlmostEqual(ta["Untagged"]["accuracy"], 1.0, places=4)


# ===========================================================================
# Task 2 — improvement() service tests
# ===========================================================================

class ImprovementServiceTest(TestCase):
    """Tests for analytics.services.improvement(test)."""

    def setUp(self):
        self.user = _make_user("improve@test.com")
        # test1 is the original; test2 is a retest (parent=test1, attempt_number=2)
        cg = ClassGroup.objects.create(user=self.user, created_by=self.user, name="CG-Improve")
        self.test1 = Test.objects.create(
            user=self.user, created_by=self.user, class_group=cg,
            title="Improve Test", subject="Sci",
            attempt_number=1,
        )
        MarkingScheme.objects.create(test=self.test1, marks_per_correct=1)
        self.test2 = Test.objects.create(
            user=self.user, created_by=self.user, class_group=cg,
            title="Improve Retest", subject="Sci",
            parent_test=self.test1, attempt_number=2,
        )
        MarkingScheme.objects.create(test=self.test2, marks_per_correct=1)
        # One question per test (just needed for sheets)
        self.q1 = Question.objects.create(test=self.test1, order_index=0, text="Q1", topic="T")
        Option.objects.create(question=self.q1, label="A", text="A", is_correct=True)
        self.q2 = Question.objects.create(test=self.test2, order_index=0, text="Q2", topic="T")
        Option.objects.create(question=self.q2, label="A", text="A", is_correct=True)

    def _make_attempt_result(self, test, question, student, score, max_score=3, seed=1):
        sheet = _make_omr_sheet(test, [question], seed=seed)
        sheet.student = student
        sheet.save()
        result = _make_result(test, student, sheet, score=score, max_score=max_score)
        return result

    def test_chain_ordered_by_attempt_number(self):
        """improvement() returns attempts ordered by attempt_number."""
        from analytics.services import improvement
        student = _make_student(self.user, "I001", "Imp Student")
        self._make_attempt_result(self.test1, self.q1, student, score=1, max_score=3, seed=800)
        self._make_attempt_result(self.test2, self.q2, student, score=2, max_score=3, seed=801)

        result = improvement(self.test2)  # call with the retest
        students_data = result["students"]
        self.assertIn("I001", students_data)
        attempts = students_data["I001"]
        attempt_nums = [a["attempt_number"] for a in attempts]
        self.assertEqual(attempt_nums, sorted(attempt_nums))

    def test_delta_vs_prev_positive_for_improvement(self):
        """
        Student goes from 1/3 (33.3%) to 2/3 (66.7%).
        delta_vs_prev for attempt 2 should be positive (~33.3%).
        """
        from analytics.services import improvement
        student = _make_student(self.user, "I002", "Delta Student")
        self._make_attempt_result(self.test1, self.q1, student, score=1, max_score=3, seed=810)
        self._make_attempt_result(self.test2, self.q2, student, score=2, max_score=3, seed=811)

        result = improvement(self.test2)
        attempts = result["students"]["I002"]
        # attempt 1: no previous → delta should be None
        # attempt 2: delta = 66.7% - 33.3% = 33.3% (positive)
        a1 = next(a for a in attempts if a["attempt_number"] == 1)
        a2 = next(a for a in attempts if a["attempt_number"] == 2)
        self.assertIsNone(a1["delta_vs_prev"])
        self.assertGreater(a2["delta_vs_prev"], 0)
        self.assertAlmostEqual(a2["delta_vs_prev"], 100.0 / 3.0, places=1)

    def test_class_average_per_attempt(self):
        """
        Two students:
          test1: student A=1/3, student B=3/3 → avg=2/3=66.7%
          test2: student A=2/3, student B=3/3 → avg=5/6=83.3%
        """
        from analytics.services import improvement
        sA = _make_student(self.user, "CA1", "Avg A")
        sB = _make_student(self.user, "CA2", "Avg B")
        self._make_attempt_result(self.test1, self.q1, sA, score=1, max_score=3, seed=820)
        self._make_attempt_result(self.test1, self.q1, sB, score=3, max_score=3, seed=821)
        self._make_attempt_result(self.test2, self.q2, sA, score=2, max_score=3, seed=822)
        self._make_attempt_result(self.test2, self.q2, sB, score=3, max_score=3, seed=823)

        result = improvement(self.test2)
        class_avg = result["class_average"]
        # class_average keyed by attempt_number
        self.assertIn(1, class_avg)
        self.assertIn(2, class_avg)
        self.assertAlmostEqual(class_avg[1], (1.0 / 3.0 + 1.0) / 2.0 * 100.0, places=1)
        self.assertAlmostEqual(class_avg[2], (2.0 / 3.0 + 1.0) / 2.0 * 100.0, places=1)

    def test_missing_attempt_handled_gracefully(self):
        """
        Student B only appears in test2, not test1.
        Their attempt list should still just have attempt 2 with delta=None.
        """
        from analytics.services import improvement
        sA = _make_student(self.user, "MH1", "Only Attempt 1")
        sB = _make_student(self.user, "MH2", "Only Attempt 2")
        self._make_attempt_result(self.test1, self.q1, sA, score=1, max_score=3, seed=830)
        self._make_attempt_result(self.test2, self.q2, sB, score=2, max_score=3, seed=831)

        result = improvement(self.test2)
        self.assertIn("MH1", result["students"])
        self.assertIn("MH2", result["students"])
        # sA: attempt 1 only
        sA_attempts = result["students"]["MH1"]
        self.assertEqual(len(sA_attempts), 1)
        self.assertEqual(sA_attempts[0]["attempt_number"], 1)
        self.assertIsNone(sA_attempts[0]["delta_vs_prev"])
        # sB: attempt 2 only
        sB_attempts = result["students"]["MH2"]
        self.assertEqual(len(sB_attempts), 1)
        self.assertEqual(sB_attempts[0]["attempt_number"], 2)
        self.assertIsNone(sB_attempts[0]["delta_vs_prev"])

    def test_improvement_called_from_root(self):
        """Calling improvement() on the ROOT test should also work correctly."""
        from analytics.services import improvement
        student = _make_student(self.user, "R001", "Root Call")
        self._make_attempt_result(self.test1, self.q1, student, score=1, max_score=3, seed=840)
        self._make_attempt_result(self.test2, self.q2, student, score=3, max_score=3, seed=841)

        result = improvement(self.test1)  # called on root
        self.assertIn("R001", result["students"])
        attempts = result["students"]["R001"]
        self.assertEqual(len(attempts), 2)

    def test_trend_key_present(self):
        """improvement() result must include a 'trend' key."""
        from analytics.services import improvement
        result = improvement(self.test1)
        self.assertIn("trend", result)

    def test_result_shape(self):
        """Top-level keys: students, class_average, trend, chain."""
        from analytics.services import improvement
        result = improvement(self.test1)
        for key in ("students", "class_average", "trend", "chain"):
            self.assertIn(key, result, f"Missing top-level key: {key!r}")


# ===========================================================================
# Task 2 — API endpoint tests
# ===========================================================================

class StudentDetailEndpointTest(TestCase):
    """GET /api/v1/analytics/test/{test_id}/student/{student_id}/ tests."""

    def setUp(self):
        self.client = APIClient()
        self.user = _make_user("ep_detail@test.com")
        self.test = _make_test(self.user)
        self.questions = _make_questions(self.test, n=3)
        self.student = _make_student(self.user, "EP001", "Endpoint Student")
        sheet = _make_omr_sheet(self.test, self.questions, seed=900)
        sheet.student = self.student
        sheet.save()
        self.result = _make_result(
            self.test, self.student, sheet, score=2, max_score=3,
            correct=2, wrong=1, blank=0,
        )
        for q_idx, q in enumerate(self.questions):
            is_correct = q_idx < 2
            _make_qr(self.result, q, q_idx, ["A"] if is_correct else ["B"], is_correct)

    def _url(self, test_id=None, student_id=None):
        test_id = test_id or self.test.id
        student_id = student_id or self.student.id
        return f"/api/v1/analytics/test/{test_id}/student/{student_id}/"

    def test_unauthenticated_401(self):
        response = self.client.get(self._url())
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_authenticated_owner_200(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url())
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_response_shape(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url())
        data = response.json()
        for key in ("score", "max_score", "pct", "correct", "wrong", "blank",
                    "per_question", "topic_accuracy"):
            self.assertIn(key, data, f"Missing key: {key!r}")

    def test_cross_tenant_test_404(self):
        other_user = _make_user("other_ep@test.com")
        other_test = _make_test(other_user)
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(test_id=other_test.id, student_id=self.student.id))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_student_not_in_test_404(self):
        """StudentResult for (test, student) must exist, otherwise 404."""
        other_student = _make_student(self.user, "EP999", "Ghost")
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(student_id=other_student.id))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)


class ImprovementEndpointTest(TestCase):
    """GET /api/v1/analytics/test/{test_id}/improvement/ tests."""

    def setUp(self):
        self.client = APIClient()
        self.user = _make_user("ep_improve@test.com")
        cg = ClassGroup.objects.create(user=self.user, created_by=self.user, name="CG-EP-Imp")
        self.test1 = Test.objects.create(
            user=self.user, created_by=self.user, class_group=cg,
            title="EP Improve Test", subject="Sci", attempt_number=1,
        )
        MarkingScheme.objects.create(test=self.test1, marks_per_correct=1)
        self.test2 = Test.objects.create(
            user=self.user, created_by=self.user, class_group=cg,
            title="EP Improve Retest", subject="Sci",
            parent_test=self.test1, attempt_number=2,
        )
        MarkingScheme.objects.create(test=self.test2, marks_per_correct=1)

    def _url(self, test_id=None):
        test_id = test_id or self.test1.id
        return f"/api/v1/analytics/test/{test_id}/improvement/"

    def test_unauthenticated_401(self):
        response = self.client.get(self._url())
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_authenticated_owner_200(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url())
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_response_shape(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url())
        data = response.json()
        for key in ("students", "class_average", "trend", "chain"):
            self.assertIn(key, data, f"Missing key: {key!r}")

    def test_cross_tenant_test_404(self):
        other_user = _make_user("other_ep_imp@test.com")
        other_test = _make_test(other_user)
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(other_test.id))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)


# ===========================================================================
# Task 3 — Export endpoint tests (CSV / Excel / PDF)
# ===========================================================================

class ExportEndpointTest(TestCase):
    """
    Tests for GET /api/v1/analytics/test/{test_id}/export/?format=csv|xlsx|pdf

    Covers:
    - Each format returns 200, non-empty body, correct Content-Type,
      Content-Disposition with "attachment".
    - CSV body contains a header row and one row per student (roll number
      appears in the body).
    - Unknown format → 400.
    - Cross-tenant request → 404.
    - Unauthenticated → 401.
    """

    def setUp(self):
        self.client = APIClient()
        self.user = _make_user("export@test.com")
        self.test = _make_test(self.user)
        self.questions = _make_questions(self.test, n=3)

        # Two students with known roll numbers
        self.student_a = _make_student(self.user, "EXP001", "Alice Export")
        self.student_b = _make_student(self.user, "EXP002", "Bob Export")

        sheet_a = _make_omr_sheet(self.test, self.questions, seed=1001)
        sheet_a.student = self.student_a
        sheet_a.save()
        self.result_a = _make_result(
            self.test, self.student_a, sheet_a,
            score=3, max_score=3, correct=3, wrong=0, blank=0,
        )
        for q_idx, q in enumerate(self.questions):
            _make_qr(self.result_a, q, q_idx, ["A"], is_correct=True)

        sheet_b = _make_omr_sheet(self.test, self.questions, seed=1002)
        sheet_b.student = self.student_b
        sheet_b.save()
        self.result_b = _make_result(
            self.test, self.student_b, sheet_b,
            score=1, max_score=3, correct=1, wrong=2, blank=0,
        )
        for q_idx, q in enumerate(self.questions):
            is_correct = q_idx == 0
            _make_qr(self.result_b, q, q_idx, ["A"] if is_correct else ["B"], is_correct)

    def _url(self, test_id=None, fmt="csv"):
        test_id = test_id or self.test.id
        return f"/api/v1/analytics/test/{test_id}/export/?output_format={fmt}"

    # ------------------------------------------------------------------ auth

    def test_unauthenticated_returns_401(self):
        response = self.client.get(self._url())
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    # ------------------------------------------------------------------ cross-tenant

    def test_cross_tenant_returns_404(self):
        other_user = _make_user("export_other@test.com")
        other_test = _make_test(other_user)
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(test_id=other_test.id))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    # ------------------------------------------------------------------ unknown format

    def test_unknown_format_returns_400(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="docx"))
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_unknown_format_error_message(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="xml"))
        data = response.json()
        self.assertIn("detail", data)

    # ------------------------------------------------------------------ default (no ?format=)

    def test_default_format_is_csv(self):
        """Omitting ?format= should default to CSV."""
        self.client.force_authenticate(user=self.user)
        url = f"/api/v1/analytics/test/{self.test.id}/export/"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("text/csv", response["Content-Type"])

    # ------------------------------------------------------------------ CSV

    def test_csv_returns_200(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="csv"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_csv_content_type(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="csv"))
        self.assertIn("text/csv", response["Content-Type"])

    def test_csv_content_disposition_attachment(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="csv"))
        cd = response.get("Content-Disposition", "")
        self.assertIn("attachment", cd)
        self.assertIn(".csv", cd)

    def test_csv_non_empty_body(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="csv"))
        self.assertGreater(len(response.content), 0)

    def test_csv_has_header_row(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="csv"))
        text = response.content.decode("utf-8")
        first_line = text.splitlines()[0]
        self.assertIn("roll_number", first_line)
        self.assertIn("score", first_line)

    def test_csv_has_one_row_per_student(self):
        """CSV should have 1 header row + 2 data rows (one per StudentResult)."""
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="csv"))
        text = response.content.decode("utf-8")
        lines = [ln for ln in text.splitlines() if ln.strip()]
        # 1 header + 2 students
        self.assertEqual(len(lines), 3)

    def test_csv_contains_roll_numbers(self):
        """Roll numbers of both students must appear in the CSV body."""
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="csv"))
        text = response.content.decode("utf-8")
        self.assertIn("EXP001", text)
        self.assertIn("EXP002", text)

    # ------------------------------------------------------------------ XLSX

    def test_xlsx_returns_200(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="xlsx"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_xlsx_content_type(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="xlsx"))
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )

    def test_xlsx_content_disposition_attachment(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="xlsx"))
        cd = response.get("Content-Disposition", "")
        self.assertIn("attachment", cd)
        self.assertIn(".xlsx", cd)

    def test_xlsx_non_empty_body(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="xlsx"))
        self.assertGreater(len(response.content), 0)

    def test_xlsx_is_valid_workbook(self):
        """The response body must be a parseable openpyxl workbook."""
        import io as _io
        from openpyxl import load_workbook

        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="xlsx"))
        wb = load_workbook(filename=_io.BytesIO(response.content))
        self.assertIn("Results", wb.sheetnames)
        self.assertIn("Summary", wb.sheetnames)

    def test_xlsx_results_sheet_has_data_rows(self):
        """Results sheet must have 1 header + 2 data rows."""
        import io as _io
        from openpyxl import load_workbook

        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="xlsx"))
        wb = load_workbook(filename=_io.BytesIO(response.content))
        ws = wb["Results"]
        # max_row includes header
        self.assertEqual(ws.max_row, 3)

    # ------------------------------------------------------------------ PDF

    def test_pdf_returns_200(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="pdf"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_pdf_content_type(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="pdf"))
        self.assertIn("application/pdf", response["Content-Type"])

    def test_pdf_content_disposition_attachment(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="pdf"))
        cd = response.get("Content-Disposition", "")
        self.assertIn("attachment", cd)
        self.assertIn(".pdf", cd)

    def test_pdf_non_empty_body(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="pdf"))
        self.assertGreater(len(response.content), 0)

    def test_pdf_magic_bytes(self):
        """A valid PDF starts with the %PDF magic bytes."""
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self._url(fmt="pdf"))
        self.assertTrue(
            response.content[:4] == b"%PDF",
            "PDF response does not start with %PDF magic bytes",
        )

    # ------------------------------------------------------------------ pure-function tests

    def test_results_csv_pure(self):
        """analytics.export.results_csv() returns correct UTF-8 CSV bytes."""
        from analytics.export import results_csv
        data = results_csv(self.test)
        self.assertIsInstance(data, bytes)
        text = data.decode("utf-8")
        lines = [ln for ln in text.splitlines() if ln.strip()]
        # header + 2 students
        self.assertEqual(len(lines), 3)
        self.assertIn("EXP001", text)
        self.assertIn("EXP002", text)

    def test_results_xlsx_pure(self):
        """analytics.export.results_xlsx() returns a parseable workbook."""
        import io as _io
        from openpyxl import load_workbook
        from analytics.export import results_xlsx

        data = results_xlsx(self.test)
        self.assertIsInstance(data, bytes)
        wb = load_workbook(filename=_io.BytesIO(data))
        self.assertIn("Results", wb.sheetnames)
        self.assertIn("Summary", wb.sheetnames)

    def test_report_pdf_pure(self):
        """analytics.export.report_pdf() returns PDF bytes."""
        from analytics.export import report_pdf

        data = report_pdf(self.test)
        self.assertIsInstance(data, bytes)
        self.assertTrue(data[:4] == b"%PDF")
