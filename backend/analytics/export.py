"""
analytics.export — produce downloadable files (CSV, Excel, PDF) for a test's results.

Public API
----------
results_csv(test)  -> bytes   (UTF-8 CSV)
results_xlsx(test) -> bytes   (Excel .xlsx)
report_pdf(test)   -> bytes   (PDF report via ReportLab)

All functions accept a `assessments.Test` instance.
"""

from __future__ import annotations

import csv
import io

from analytics.services import test_summary


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

_CSV_HEADERS = [
    "roll_number",
    "full_name",
    "score",
    "max_score",
    "correct",
    "wrong",
    "blank",
    "needs_review",
]


def results_csv(test) -> bytes:
    """
    Return a UTF-8-encoded CSV bytes object with one header row and one data
    row per StudentResult for *test*.

    Columns: roll_number, full_name, score, max_score, correct, wrong, blank,
             needs_review
    """
    from results.models import StudentResult

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_CSV_HEADERS)

    qs = (
        StudentResult.objects
        .filter(test=test)
        .select_related("student")
        .order_by("student__roll_number", "id")
    )
    for sr in qs:
        student = sr.student
        roll = student.roll_number if student else ""
        name = (student.full_name or "") if student else ""
        writer.writerow([
            roll,
            name,
            str(sr.score),
            str(sr.max_score),
            sr.correct_count,
            sr.wrong_count,
            sr.blank_count,
            "yes" if sr.needs_review else "no",
        ])

    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# Excel (openpyxl)
# ---------------------------------------------------------------------------

def results_xlsx(test) -> bytes:
    """
    Return an Excel .xlsx bytes object with two sheets:

    "Results"  — same columns as the CSV, with a bold header row.
    "Summary"  — key stats from test_summary(): n_students, average, median,
                 max, min, max_score, and the score distribution.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from results.models import StudentResult

    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1 — Results
    # ------------------------------------------------------------------
    ws_results = wb.active
    ws_results.title = "Results"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="366092")

    for col_idx, header in enumerate(_CSV_HEADERS, start=1):
        cell = ws_results.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    qs = (
        StudentResult.objects
        .filter(test=test)
        .select_related("student")
        .order_by("student__roll_number", "id")
    )
    for row_idx, sr in enumerate(qs, start=2):
        student = sr.student
        roll = student.roll_number if student else ""
        name = (student.full_name or "") if student else ""
        ws_results.append([
            roll,
            name,
            float(sr.score),
            float(sr.max_score),
            sr.correct_count,
            sr.wrong_count,
            sr.blank_count,
            "yes" if sr.needs_review else "no",
        ])

    # Auto-size columns (rough heuristic)
    for col in ws_results.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws_results.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ------------------------------------------------------------------
    # Sheet 2 — Summary
    # ------------------------------------------------------------------
    ws_summary = wb.create_sheet(title="Summary")

    summary = test_summary(test)

    summary_rows = [
        ("Test", summary["test"]["title"]),
        ("Subject", summary["test"].get("subject", "")),
        ("Attempt", summary["test"]["attempt_number"]),
        ("", ""),
        ("Students", summary["n_students"]),
        ("Graded", summary["graded"]),
        ("Needs Review", summary["needs_review_count"]),
        ("", ""),
        ("Average Score", summary["average"] if summary["average"] is not None else "N/A"),
        ("Median Score", summary["median"] if summary["median"] is not None else "N/A"),
        ("Max Score", summary["max"] if summary["max"] is not None else "N/A"),
        ("Min Score", summary["min"] if summary["min"] is not None else "N/A"),
        ("Max Possible", summary["max_score"] if summary["max_score"] is not None else "N/A"),
        ("", ""),
        ("Score Distribution", ""),
    ]
    for label, value in summary_rows:
        ws_summary.append([label, value])

    # Distribution sub-table
    ws_summary.append(["Bucket", "Count"])
    for bucket_row in summary["distribution"]:
        ws_summary.append([bucket_row["bucket"], bucket_row["count"]])

    # Bold the header labels
    label_font = Font(bold=True)
    for row in ws_summary.iter_rows():
        if row[0].value and row[1].value == "":
            row[0].font = label_font
        # Bold the section headers
        if row[0].value in ("Score Distribution", "Bucket"):
            row[0].font = label_font

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30

    # ------------------------------------------------------------------
    # Serialise
    # ------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF (ReportLab)
# ---------------------------------------------------------------------------

def report_pdf(test) -> bytes:
    """
    Return a printable PDF report as bytes.

    Sections
    --------
    1. Title block  — test title, subject, attempt number.
    2. Summary stats — n_students, average, median, max, min, max_possible.
    3. Score distribution table.
    4. Toppers list (up to 5).
    5. Hardest questions list (up to 10).
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    summary = test_summary(test)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title=f"Results Report — {summary['test']['title']}",
    )

    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h2 = styles["Heading2"]
    normal = styles["Normal"]
    small = ParagraphStyle("small", parent=normal, fontSize=9)

    story = []

    # ------------------------------------------------------------------
    # 1. Title block
    # ------------------------------------------------------------------
    story.append(Paragraph("OMRFlow — Results Report", h1))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(f"<b>Test:</b> {summary['test']['title']}", normal))
    story.append(Paragraph(f"<b>Subject:</b> {summary['test'].get('subject', '')}", normal))
    story.append(Paragraph(f"<b>Attempt:</b> {summary['test']['attempt_number']}", normal))
    story.append(Spacer(1, 0.5 * cm))

    # ------------------------------------------------------------------
    # 2. Summary stats
    # ------------------------------------------------------------------
    story.append(Paragraph("Summary Statistics", h2))

    def _fmt(v):
        if v is None:
            return "N/A"
        if isinstance(v, float):
            return f"{v:.2f}"
        return str(v)

    stats_data = [
        ["Metric", "Value"],
        ["Students", _fmt(summary["n_students"])],
        ["Graded", _fmt(summary["graded"])],
        ["Needs Review", _fmt(summary["needs_review_count"])],
        ["Average Score", _fmt(summary["average"])],
        ["Median Score", _fmt(summary["median"])],
        ["Max Score (achieved)", _fmt(summary["max"])],
        ["Min Score (achieved)", _fmt(summary["min"])],
        ["Max Possible Score", _fmt(summary["max_score"])],
    ]

    stats_table = Table(stats_data, colWidths=[9 * cm, 6 * cm])
    stats_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF3F8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 0.5 * cm))

    # ------------------------------------------------------------------
    # 3. Score distribution table
    # ------------------------------------------------------------------
    story.append(Paragraph("Score Distribution", h2))

    dist_data = [["Bucket", "Count"]]
    for bucket in summary["distribution"]:
        dist_data.append([bucket["bucket"], str(bucket["count"])])

    dist_table = Table(dist_data, colWidths=[7 * cm, 4 * cm])
    dist_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF3F8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(dist_table)
    story.append(Spacer(1, 0.5 * cm))

    # ------------------------------------------------------------------
    # 4. Toppers
    # ------------------------------------------------------------------
    if summary["toppers"]:
        story.append(Paragraph("Top Performers", h2))
        topper_data = [["Rank", "Roll No.", "Name", "Score", "Max"]]
        for rank, t in enumerate(summary["toppers"], start=1):
            topper_data.append([
                str(rank),
                t["student"]["roll"],
                t["student"]["name"],
                _fmt(t["score"]),
                _fmt(t["max_score"]),
            ])
        topper_table = Table(
            topper_data,
            colWidths=[1.5 * cm, 3 * cm, 6 * cm, 2.5 * cm, 2.5 * cm],
        )
        topper_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF3F8")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(topper_table)
        story.append(Spacer(1, 0.5 * cm))

    # ------------------------------------------------------------------
    # 5. Hardest questions
    # ------------------------------------------------------------------
    hq = summary["hardest_questions"][:10]
    if hq:
        story.append(Paragraph("Hardest Questions (by wrong-answer rate)", h2))
        hq_data = [["#", "Question", "Wrong Rate", "Responses"]]
        for rank, q in enumerate(hq, start=1):
            # Truncate long question text
            text = q["text"]
            if len(text) > 60:
                text = text[:57] + "..."
            hq_data.append([
                str(rank),
                text,
                f"{q['wrong_rate'] * 100:.1f}%",
                str(q["n"]),
            ])
        hq_table = Table(
            hq_data,
            colWidths=[1 * cm, 10 * cm, 3 * cm, 2.5 * cm],
        )
        hq_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF3F8")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(hq_table)

    # ------------------------------------------------------------------
    # Build
    # ------------------------------------------------------------------
    doc.build(story)
    return buf.getvalue()
